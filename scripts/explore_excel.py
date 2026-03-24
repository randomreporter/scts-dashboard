import openpyxl

files = {
    'PRE_MID': r'D:\AI\AI Projects\Anti gravity\SCTS\CBSE KA TN 6-9 PRE MID TERM Marks Entry Format (1) (8).xlsx',
    'MID_TERM': r'D:\AI\AI Projects\Anti gravity\SCTS\CBSE KA TN MID TERM Marks Entry Format (5).xlsx',
    'POST_MID': r'D:\AI\AI Projects\Anti gravity\SCTS\CBSE KA TN POST MID Marks Entry Format (3).xlsx',
    'ANNUAL': r'D:\AI\AI Projects\Anti gravity\SCTS\CBSE KA TN ANNUAL EXAM Marks Entry Format (1).xlsx',
}

for label, fpath in files.items():
    wb = openpyxl.load_workbook(fpath, data_only=True)
    print(f'\n{"="*60}')
    print(f'{label}: {fpath.split(chr(92))[-1]}')
    print(f'Sheets: {wb.sheetnames}')
    for sname in wb.sheetnames:
        ws = wb[sname]
        print(f'\n  -- Sheet: {sname} (rows={ws.max_row}, cols={ws.max_column}) --')
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=6, values_only=True)):
            print(f'  Row {i+1}: {row}')
